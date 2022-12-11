import openpyxl
import pprint
from typing import Union
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('Price-one.xlsx')
sheet1 = wb.active

wb2 = openpyxl.load_workbook('Price-two.xlsx')
sheet2 = wb2.active


def intersection(lst1, lst2):
    lst3 = [value for value in lst1 if value in lst2]
    return lst3


def getProductList(sheet):
    productList = []
    for i in range(2, sheet.max_row+1):
        productList.append(sheet['A'+str(i)].value)
    return productList


def getMissedProductList(productList1, productList2):
    missedProductList = intersection(productList1, productList2)
    return missedProductList

# def getProducts():
#     missedProductList=getMissedProductList(sheet1,sheet2)
#     for i in range(2,sheet1.max_row+1):
#         if sheet1['A'+str(i)].value in missedProductList:
#             sheet1['A'+str(i)].value=''

#     for i in range(2,sheet2.max_row+1):
#         if sheet2['A'+str(i)].value in missedProductList:
#             sheet2['A'+str(i)].value=''

#     wb.save('Price-one.xlsx')
#     wb2.save('Price-two.xlsx')


def getProducts(row_min, row_max, col_min, col_max, sheet):
    Products = {}
    for row in range(row_min, row_max):
        temp_obj = {}
        for column in range(col_min, col_max):
            cell = sheet.cell(row, column)
            if cell.value != None:
                # print(sheet.cell(1, column).value, ": ", cell.value)
                temp_obj.setdefault(sheet.cell(1, column).value, cell.value)
        if temp_obj != []:
            Products.setdefault(sheet.cell(row, 1).value, temp_obj)
    return Products


def getProductParameters(sheet):
    TempList = []
    for row in range(1, 100):
        if sheet.cell(row, 1).value == "Product":
            for column in range(2, 100):
                cell = sheet.cell(row, column)
                if cell.value != None:
                    TempList.append(cell.value)
    return TempList


# getting product lists from both sheets
productListOne = getProductList(sheet1)
productListTwo = getProductList(sheet2)

# do not need this operations
# missedInFirstMarket=getMissedProductList(productListOne,productListTwo)
# missedInSecondMarket=getMissedProductList(productListTwo,productListOne)

ProductsOne = getProducts(2, sheet1.max_row+1, 2, sheet1.max_column+1, sheet1)
ProductsTwo = getProducts(2, sheet2.max_row+1, 2, sheet2.max_column+1, sheet2)

getCrossedParams = intersection(
    getProductParameters(sheet1), getProductParameters(sheet2))

print("Parameters: ", getCrossedParams)
print("Products: ", intersection(productListOne, productListTwo))

comparisonParam = str(input('Parameter to compare products: '))
productToCompare = str(input('Product: '))

chosenParams = []

for param in getCrossedParams:
    if param in comparisonParam:
        chosenParams.append(param)

chosenProducts = []

for product in intersection(productListOne, productListTwo):
    if product in productToCompare:
        chosenProducts.append(product)

for param in chosenParams:
    print("----------------------------------------------------")
    print("Parameter: ", param)

    for product in chosenProducts:
        firstVal = ProductsOne.get(product).get(param)
        secondVal = ProductsTwo.get(product).get(param)

        print("---------------------------------")
        print("Product: ", product)
        print("In first market: ", firstVal)
        print("In second market: ", secondVal)

        if param == "Price":
            if firstVal < secondVal:
                print("Cheaper in the first market")
            elif firstVal > secondVal:
                print("Cheaper in the second market")
        elif (type(firstVal) == int and type(secondVal) == int) or (type(firstVal) == float and type(secondVal) == float):
            if firstVal < secondVal:
                print(param, "is greater in the first market")
            elif firstVal > secondVal:
                print(param, "is greater in the second market")
            else:
                print(param, "is equal in both markets")
        elif type(firstVal) == str:
            equal = (ProductsOne.get(product).get(param) == (secondVal))
            print("Values are equal" if equal else "Values are not equal")
